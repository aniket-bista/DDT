from openpyxl import *


def getRowCount(file, sheetname):
    workbook = load_workbook(file)
    sheet = workbook[f'{sheetname}']
    return sheet.max_row

def getColumnCount(file, sheetname):
    workbook = load_workbook(file)
    sheet = workbook[f'{sheetname}']
    return sheet.max_column

def readData(file, sheetname, rows, cols):
    workbook = load_workbook(file)
    sheet = workbook[f'{sheetname}']
    return sheet.cell(row=rows, column=cols).value

def writeData(file, sheetname, rows, cols, data):
    workbook = load_workbook(file)
    sheet = workbook[f'{sheetname}']
    sheet.cell(row=rows, column=cols).value = data
    workbook.save(file)

